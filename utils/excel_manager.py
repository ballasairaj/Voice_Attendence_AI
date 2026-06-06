from openpyxl import Workbook, load_workbook
import os


EXCEL_FILE = "reports/attendance.xlsx"


def create_excel_file():

    if not os.path.exists(EXCEL_FILE):

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = "Attendance"

        sheet.append([
            "Student Name",
            "Class",
            "Section",
            "Status",
            "Date",
            "Time"
        ])

        workbook.save(EXCEL_FILE)

        print("Excel file created")


def update_excel(
        student_name,
        class_name,
        section,
        status,
        date,
        time
):

    workbook = load_workbook(EXCEL_FILE)

    sheet = workbook.active

    sheet.append([
        student_name,
        class_name,
        section,
        status,
        date,
        time
    ])

    workbook.save(EXCEL_FILE)

    print("Excel updated successfully")